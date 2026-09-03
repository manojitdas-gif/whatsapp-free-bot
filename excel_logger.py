"""
excel_logger.py — Exact 9-Column Customer Lead Logger.

EXACT COLUMNS:
  Col 1: First Contact Date (IST)
  Col 2: Last Contact Date (IST)
  Col 3: Contact Person Name
  Col 4: WhatsApp Number
  Col 5: Email ID
  Col 6: Company / Business Name
  Col 7: GST Number
  Col 8: Complete Address
  Col 9: Requirements  ← product details, quantities, specifications

RULES:
  - ONE ROW per customer — updated in place, never duplicated.
  - Each update ONLY fills blank cells or APPENDS to Requirements.
  - Truly blank cells when no data found (no dashes, no placeholders).
  - Data extracted from ALL formats: text, photos (OCR), PDFs, Word, Excel, CSV.
"""

import os
import sys
import re
import csv
import json
import time
from datetime import datetime, timezone, timedelta
from threading import Lock

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(errors='replace')
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE_PATH, BACKUP_EXCEL_PATH, SHARED_EXCEL_PATH

IST = timezone(timedelta(hours=5, minutes=30))
_excel_lock = Lock()

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
QUEUE_FILE = os.path.join(BASE_DIR, "data", "pending_lead_queue.json")
CSV_PATH   = os.path.join(os.path.expanduser("~"), "Desktop", "WhatsApp_Leads_Live.csv")

# ── STYLING ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1B5E20")          # Dark green
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT   = Font(name="Calibri", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F1F8E9")           # Very light green for alternate rows

THIN_BORDER = Border(
    left  =Side(style="thin", color="C8E6C9"),
    right =Side(style="thin", color="C8E6C9"),
    top   =Side(style="thin", color="C8E6C9"),
    bottom=Side(style="thin", color="C8E6C9"),
)

# ── EXACT 9 COLUMNS ────────────────────────────────────────────────────────────
HEADERS = [
    "First Contact Date (IST)",
    "Last Contact Date (IST)",
    "Contact Person Name",
    "WhatsApp Number",
    "Email ID",
    "Company / Business Name",
    "GST Number",
    "Complete Address",
    "Requirements",
]
COL_WIDTHS = [22, 22, 24, 18, 30, 32, 20, 50, 70]

# Column index constants (1-based)
C_FIRST_CONTACT = 1
C_LAST_CONTACT  = 2
C_NAME          = 3
C_PHONE         = 4
C_EMAIL         = 5
C_COMPANY       = 6
C_GST           = 7
C_ADDRESS       = 8
C_REQUIREMENTS  = 9


# ──────────────────────────────────────────────────────────────────────────────
# PHONE FORMATTER
# ──────────────────────────────────────────────────────────────────────────────

def format_phone_display(raw_phone: str) -> str:
    p = re.sub(r'[^\d]', '', str(raw_phone or ""))
    if p.startswith("91") and len(p) == 12:
        return f"+91 {p[2:7]} {p[7:]}"
    if len(p) == 10:
        return f"+91 {p[:5]} {p[5:]}"
    return f"+{p}" if p else "Unknown"


# ──────────────────────────────────────────────────────────────────────────────
# ENTITY EXTRACTOR — extract all 6 data fields from any text with high accuracy
# ──────────────────────────────────────────────────────────────────────────────

# Patterns
_EMAIL_RE   = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,7}\b')
_GST_RE     = re.compile(r'\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]\b', re.IGNORECASE)
_PHONE_RE   = re.compile(r'\b(?:\+91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}\b')
_PIN_RE     = re.compile(r'\b[1-9]\d{5}\b')
_URL_RE     = re.compile(r'https?://\S+|www\.\S+', re.IGNORECASE)

_ADDR_KW = re.compile(
    r'\b(road|rd\b|street|st\b|lane|ln\b|gali|bazar|bazaar|nagar|colony|block|sector|'
    r'floor|flat|shop|unit|near|opp|opposite|behind|beside|above|below|next to|'
    r'dist|district|taluk|tehsil|mandal|'
    r'kolkata|calcutta|mumbai|bombay|delhi|new delhi|chennai|madras|bangalore|bengaluru|'
    r'hyderabad|pune|ahmedabad|surat|jaipur|lucknow|kanpur|nagpur|indore|bhopal|'
    r'howrah|siliguri|durgapur|asansol|rajkot|vadodara|coimbatore|kochi|'
    r'west bengal|maharashtra|gujarat|uttar pradesh|rajasthan|karnataka|tamil nadu|'
    r'telangana|andhra|kerala|bihar|jharkhand|odisha|assam|punjab|haryana|'
    r'p\.?o\.?\b|p\.?s\.?\b|via\b)\b',
    re.IGNORECASE
)

_BIZ_KW = re.compile(
    r'\b(enterprise[s]?|pvt\.?\s*ltd\.?|limited|llp|llc|'
    r'trader[s]?|trading|manufacturer[s]?|supplier[s]?|distributor[s]?|'
    r'industries|industry|industrial|packaging|packers|'
    r'corporation|corp\b|company|co\b|group|associates|agency|agencies|'
    r'shop|store|mart|house|hub|center|centre|works|workshop|'
    r'solutions|services|systems|tech|technologies)\b',
    re.IGNORECASE
)

_CHATTER_WORDS = {
    "hi", "hello", "hey", "hii", "helo", "ok", "okay", "k",
    "yes", "no", "sure", "fine", "good", "done", "noted",
    "thanks", "thank you", "thank you sir", "thanks sir",
    "please", "kindly", "send", "reply", "call me", "contact",
    "start", "restart", "haan", "ha", "ji", "namaste",
    "good morning", "good afternoon", "good evening", "good night",
}


def _is_chatter(text: str) -> bool:
    return text.strip().lower() in _CHATTER_WORDS


def extract_entities(text: str, profile_name: str = "") -> dict:
    """
    Extract all 6 entity fields from customer message text:
    name, email, company, gst, address, phone (for validation only).
    Returns dict with keys: contact_name, email, company, gst, address.
    All values are strings — empty string if not found.
    """
    text = str(text or "").strip()
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # ── EMAIL ──────────────────────────────────────────────────────────────────
    email_m = _EMAIL_RE.search(text)
    email = email_m.group(0).lower() if email_m else ""

    # ── GST NUMBER ─────────────────────────────────────────────────────────────
    gst_m = _GST_RE.search(text)
    gst = gst_m.group(0).upper() if gst_m else ""

    # ── CONTACT PERSON NAME ────────────────────────────────────────────────────
    contact_name = ""

    # Pattern 1: explicit label
    name_label_m = re.search(
        r'(?:contact\s*(?:person|name)|person\s*name|name)\s*[:\-–=]\s*([A-Za-z][A-Za-z\s\.]{2,35})',
        text, re.IGNORECASE
    )
    if name_label_m:
        candidate = name_label_m.group(1).strip().rstrip('.,;')
        if not _BIZ_KW.search(candidate) and len(candidate.split()) <= 5:
            contact_name = candidate.title()

    # Pattern 2: "My name is ..."
    if not contact_name:
        name_is_m = re.search(r'my\s+name\s+is\s+([A-Za-z][A-Za-z\s\.]{2,30})', text, re.IGNORECASE)
        if name_is_m:
            contact_name = name_is_m.group(1).strip().rstrip('.,;').title()

    # Pattern 3: WhatsApp profile name (fallback)
    if not contact_name and profile_name:
        pn = profile_name.strip()
        if pn.lower() not in ("customer", "unknown", "business account") and len(pn) >= 2:
            contact_name = pn.title()

    # ── COMPANY / BUSINESS NAME ────────────────────────────────────────────────
    company = ""

    # Pattern 1: explicit label
    comp_label_m = re.search(
        r'(?:company|business|firm|shop|organisation|organization|org)\s*(?:name)?\s*[:\-–=]\s*([^\n,;]{3,60})',
        text, re.IGNORECASE
    )
    if comp_label_m:
        company = comp_label_m.group(1).strip().rstrip('.,;')

    # Pattern 2: line containing business keywords
    if not company:
        for line in lines:
            if _BIZ_KW.search(line):
                # Reject if it's an address line or chatter
                if not _ADDR_KW.search(line) and not _is_chatter(line):
                    if not gst or gst.lower() not in line.lower():
                        if 3 <= len(line) <= 70:
                            company = line.strip().rstrip('.,;')
                            break

    # ── COMPLETE ADDRESS ───────────────────────────────────────────────────────
    address_lines = []
    for line in lines:
        if not line:
            continue
        # Skip if it's the GST line or email line
        if gst and gst.lower() in line.lower():
            continue
        if email and email in line.lower():
            continue
        if _URL_RE.search(line):
            continue
        if _is_chatter(line):
            continue
        # Skip if it's the company line
        if company and line.lower().strip() == company.lower().strip():
            continue
        # Skip if it's the contact name line alone
        if contact_name and line.lower().strip() == contact_name.lower().strip():
            continue
        # Skip phone-only lines
        if _PHONE_RE.match(line) and len(line) < 16:
            continue

        # Accept if it has address keywords OR a pin code
        if _ADDR_KW.search(line) or _PIN_RE.search(line):
            cleaned = re.sub(
                r'^(?:address|location|office\s*address|addr|add)\s*[:\-–=]\s*',
                '', line, flags=re.IGNORECASE
            ).strip()
            if cleaned and cleaned not in address_lines:
                address_lines.append(cleaned)

    # If no keyword match but text is clearly multi-line address-like (has pin code)
    if not address_lines and _PIN_RE.search(text):
        # Grab up to 3 lines around the pin code line
        for i, line in enumerate(lines):
            if _PIN_RE.search(line):
                start = max(0, i - 2)
                end = min(len(lines), i + 2)
                address_lines = [l for l in lines[start:end] if not _is_chatter(l)]
                break

    address = ", ".join(address_lines) if address_lines else ""

    # ── FALLBACK: for business step, if company still not found, try first
    #    non-chatter, non-address, non-gst, non-email, non-name line ────────────
    if not company:
        for line in lines:
            if _is_chatter(line):
                continue
            if gst and gst.lower() in line.lower():
                continue
            if email and email in line.lower():
                continue
            if _ADDR_KW.search(line) or _PIN_RE.search(line):
                continue
            if _PHONE_RE.match(line) and len(line) < 16:
                continue
            if contact_name and line.lower().strip() == contact_name.lower().strip():
                continue
            if 3 <= len(line) <= 70 and not re.search(r'\b(need|want|require|send|quote|price|rate|product|item)\b', line, re.IGNORECASE):
                company = line.strip().rstrip('.,;')
                break

    return {
        "contact_name": contact_name,
        "email":        email,
        "company":      company,
        "gst":          gst,
        "address":      address,
    }


# ──────────────────────────────────────────────────────────────────────────────
# WORKBOOK HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _get_or_create_workbook(file_path: str) -> openpyxl.Workbook:
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    wb = None
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception:
            wb = None

    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Leads"
    else:
        ws = wb["Customer Leads"] if "Customer Leads" in wb.sheetnames else wb.active

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Always ensure row 1 has all exact 9 headers and correct formatting
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    return wb


def _apply_lead_to_workbook(file_path: str, lead: dict) -> bool:
    """Write or update a customer lead row — 9 columns."""
    wb = _get_or_create_workbook(file_path)
    ws = wb["Customer Leads"] if "Customer Leads" in wb.sheetnames else wb.active

    phone        = lead.get("phone", "")
    ts           = lead.get("timestamp", "")
    name         = lead.get("contact_name", "")
    email        = lead.get("email", "")
    company      = lead.get("company", "")
    gst          = lead.get("gst", "")
    address      = lead.get("address", "")
    requirements = lead.get("requirements", "")

    align_c = Alignment(horizontal="center", vertical="center")
    align_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Find existing row by WhatsApp Number (column 4)
    match_row   = None
    first_empty = None
    row_count   = 0

    for r in range(2, ws.max_row + 2):
        cv = ws.cell(row=r, column=C_PHONE).value
        if cv:
            row_count += 1
            if str(cv).strip() == str(phone).strip():
                match_row = r
                break
        elif first_empty is None:
            first_empty = r

    def _fill_if_empty(row, col, new_val):
        """Only write if cell is currently blank and we have a value."""
        if new_val:
            existing = ws.cell(row=row, column=col).value
            if not existing or str(existing).strip() == "":
                ws.cell(row=row, column=col, value=new_val)

    def _append_requirements(row, new_req):
        """Append new requirement text to existing — never overwrite."""
        if not new_req:
            return
        existing = str(ws.cell(row=row, column=C_REQUIREMENTS).value or "").strip()
        if not existing:
            ws.cell(row=row, column=C_REQUIREMENTS, value=new_req)
        else:
            # Only append if it contains genuinely new info (not already in the cell)
            if new_req.strip().lower() not in existing.lower():
                ws.cell(row=row, column=C_REQUIREMENTS, value=existing + "\n" + new_req)

    if match_row:
        # Always update Last Contact Date
        ws.cell(row=match_row, column=C_LAST_CONTACT, value=ts)
        # Fill other columns only if currently empty
        _fill_if_empty(match_row, C_NAME,    name)
        _fill_if_empty(match_row, C_EMAIL,   email)
        _fill_if_empty(match_row, C_COMPANY, company)
        _fill_if_empty(match_row, C_GST,     gst)
        _fill_if_empty(match_row, C_ADDRESS, address)
        # Requirements: always APPEND new product info
        _append_requirements(match_row, requirements)

    else:
        new_row = first_empty if first_empty else (ws.max_row + 1)
        row_vals = [
            ts,           # C1: First Contact Date
            ts,           # C2: Last Contact Date
            name,         # C3: Contact Person Name
            phone,        # C4: WhatsApp Number
            email,        # C5: Email ID
            company,      # C6: Company / Business Name
            gst,          # C7: GST Number
            address,      # C8: Complete Address
            requirements, # C9: Requirements
        ]
        is_alt = (new_row % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=new_row, column=col_idx, value=val if val else None)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = ALT_FILL
            cell.alignment = (
                align_c if col_idx in (C_FIRST_CONTACT, C_LAST_CONTACT, C_PHONE, C_GST)
                else align_l
            )
        ws.row_dimensions[new_row].height = 28

    wb.save(file_path)
    return True


# ──────────────────────────────────────────────────────────────────────────────
# QUEUE (for when Excel file is open/locked)
# ──────────────────────────────────────────────────────────────────────────────

def _load_queue() -> list:
    if os.path.exists(QUEUE_FILE):
        try:
            with open(QUEUE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def _save_queue(queue: list) -> None:
    os.makedirs(os.path.dirname(QUEUE_FILE) or ".", exist_ok=True)
    try:
        with open(QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump(queue, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def flush_pending_excel_queue():
    with _excel_lock:
        queue = _load_queue()
        if not queue:
            return
        unflushed = []
        flushed = 0
        for item in queue:
            dest, lead = (item[0], item[1]) if isinstance(item, (list, tuple)) else (EXCEL_FILE_PATH, item)
            try:
                _apply_lead_to_workbook(dest, lead)
                flushed += 1
            except PermissionError:
                unflushed.append((dest, lead))
            except Exception as e:
                print(f"[EXCEL QUEUE ERROR] {e}")
        if flushed:
            print(f"[EXCEL SYNC] Flushed {flushed} queued lead(s) to Desktop files.")
        _save_queue(unflushed)


# ──────────────────────────────────────────────────────────────────────────────
# CSV SYNC
# ──────────────────────────────────────────────────────────────────────────────

def _update_live_csv(lead: dict) -> None:
    try:
        rows = []
        if os.path.exists(CSV_PATH):
            with open(CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
        if not rows:
            rows = [HEADERS]

        phone = lead.get("phone", "")
        match_idx = None
        for i in range(1, len(rows)):
            if len(rows[i]) > 3 and rows[i][3] == phone:
                match_idx = i
                break

        def _safe_fill(rows, idx, col, val):
            while len(rows[idx]) <= col:
                rows[idx].append("")
            if val and not rows[idx][col]:
                rows[idx][col] = val

        def _safe_append(rows, idx, col, val):
            while len(rows[idx]) <= col:
                rows[idx].append("")
            if val:
                existing = rows[idx][col]
                if not existing:
                    rows[idx][col] = val
                elif val.strip().lower() not in existing.lower():
                    rows[idx][col] = existing + " | " + val

        if match_idx is not None:
            rows[match_idx][1] = lead.get("timestamp", "")  # Last contact
            _safe_fill(rows, match_idx, 2, lead.get("contact_name"))
            _safe_fill(rows, match_idx, 4, lead.get("email"))
            _safe_fill(rows, match_idx, 5, lead.get("company"))
            _safe_fill(rows, match_idx, 6, lead.get("gst"))
            _safe_fill(rows, match_idx, 7, lead.get("address"))
            _safe_append(rows, match_idx, 8, lead.get("requirements"))  # Append new reqs
        else:
            rows.append([
                lead.get("timestamp", ""),
                lead.get("timestamp", ""),
                lead.get("contact_name") or "",
                phone,
                lead.get("email") or "",
                lead.get("company") or "",
                lead.get("gst") or "",
                lead.get("address") or "",
                lead.get("requirements") or "",
            ])

        with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)
    except Exception as e:
        print(f"[CSV ERROR] {e}")


# ──────────────────────────────────────────────────────────────────────────────
# PRIMARY PUBLIC API
# ──────────────────────────────────────────────────────────────────────────────

def log_customer_lead(
    sender_phone: str,
    sender_name: str,
    message_text: str = "",
    ocr_text: str = "",
    requirements: str = "",
) -> None:
    """
    Log customer lead to all Excel/CSV files (9 columns).
    Extracts entity fields from message_text + ocr_text.
    Requirements are saved separately (passed explicitly from engine).

    Args:
        sender_phone:  raw phone number string
        sender_name:   WhatsApp profile display name
        message_text:  all customer text messages combined
        ocr_text:      text extracted from customer images/documents/files
        requirements:  product requirements already parsed by the engine
    """
    now_ist = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    formatted_phone = format_phone_display(sender_phone)

    # Combine all text for entity extraction
    combined_text = "\n".join(filter(None, [message_text.strip(), ocr_text.strip()]))
    extracted = extract_entities(combined_text, profile_name=sender_name)

    lead = {
        "timestamp":    now_ist,
        "contact_name": extracted["contact_name"],
        "phone":        formatted_phone,
        "email":        extracted["email"],
        "company":      extracted["company"],
        "gst":          extracted["gst"],
        "address":      extracted["address"],
        "requirements": requirements,   # Pre-parsed product requirements
    }

    print(f"[LEAD] {formatted_phone} | Name: '{extracted['contact_name']}' | "
          f"Co: '{extracted['company']}' | GST: '{extracted['gst']}' | "
          f"Req: '{(requirements or '')[:50]}...'", flush=True)

    # Cloud sync
    try:
        from cloud_sync import push_lead_to_cloud
        push_lead_to_cloud({
            "first_contact": now_ist, "last_contact": now_ist,
            "name": extracted["contact_name"], "phone": formatted_phone,
            "email": extracted["email"], "company": extracted["company"],
            "gst": extracted["gst"], "address": extracted["address"],
            "requirements": requirements,
        })
    except Exception:
        pass

    with _excel_lock:
        _update_live_csv(lead)

        try:
            _apply_lead_to_workbook(BACKUP_EXCEL_PATH, lead)
        except Exception as e:
            print(f"[EXCEL BACKUP ERROR] {e}")

        for dest in [EXCEL_FILE_PATH, SHARED_EXCEL_PATH]:
            try:
                _apply_lead_to_workbook(dest, lead)
                print(f"[EXCEL] ✅ {os.path.basename(dest)}: {extracted['contact_name'] or sender_name} ({formatted_phone})")
            except PermissionError:
                queue = _load_queue()
                queue.append((dest, lead))
                _save_queue(queue)
                print(f"[EXCEL] ⚠ {os.path.basename(dest)} is open — queued.")
            except Exception as e:
                queue = _load_queue()
                queue.append((dest, lead))
                _save_queue(queue)
                print(f"[EXCEL QUEUE] {e}")
