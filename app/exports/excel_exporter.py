"""
excel_exporter.py — Exact 9-Column Excel & CSV Exporter and Synchronizer.

EXACT 9 COLUMNS:
  1. First Contact Date (IST)
  2. Last Contact Date (IST)
  3. Contact Person Name
  4. WhatsApp Number
  5. Email ID
  6. Company / Business Name
  7. GST Number
  8. Complete Address
  9. Customer Requirements Details
"""

import os
import csv
import threading
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings

IST = timezone(timedelta(hours=5, minutes=30))
_excel_lock = threading.Lock()

HEADERS = [
    "First Contact Date (IST)",
    "Last Contact Date (IST)",
    "Contact Person Name",
    "WhatsApp Number",
    "Email ID",
    "Company / Business Name",
    "GST Number",
    "Complete Address",
    "Customer Requirements Details",
]

COL_WIDTHS = [22, 22, 24, 18, 30, 32, 20, 50, 70]

HEADER_FILL = PatternFill("solid", fgColor="1B5E20")  # Dark green
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT   = Font(name="Calibri", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F1F8E9")

THIN_BORDER = Border(
    left  =Side(style="thin", color="C8E6C9"),
    right =Side(style="thin", color="C8E6C9"),
    top   =Side(style="thin", color="C8E6C9"),
    bottom=Side(style="thin", color="C8E6C9"),
)

def format_ist_timestamp(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        return dt.split()[0]
    # If naive, treat as UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST).strftime("%Y-%m-%d")

def format_phone_display(raw_phone: str) -> str:
    p = "".join(filter(str.isdigit, str(raw_phone or "")))
    if p.startswith("91") and len(p) == 12:
        return f"+91 {p[2:7]} {p[7:]}"
    if len(p) == 10:
        return f"+91 {p[:5]} {p[5:]}"
    return f"+{p}" if p else ""

def sync_customer_to_excel(customer) -> None:
    """
    Safely writes/updates a single customer master row in the 9-column Excel file.
    Updates in-place by WhatsApp Number matching, never duplicates.
    """
    with _excel_lock:
        for file_path in [settings.SHARED_EXCEL_PATH, settings.EXCEL_EXPORT_PATH]:
            try:
                _write_to_workbook(file_path, customer)
            except PermissionError:
                if file_path == settings.EXCEL_EXPORT_PATH:
                    queue_pending_customer(customer)
            except Exception as e:
                print(f"[EXCEL SYNC ERROR] {e}")

        # Update Live CSV
        try:
            _write_to_csv(settings.CSV_EXPORT_PATH, customer)
        except Exception as e:
            print(f"[CSV SYNC ERROR] {e}")

def _get_or_create_workbook(file_path: str) -> openpyxl.Workbook:
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    wb = None
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception:
            wb = None

    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Leads"
    else:
        ws = wb["Customer Leads"] if "Customer Leads" in wb.sheetnames else wb.active

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Trim extra columns if any exist beyond 9
    if ws.max_column > len(HEADERS):
        ws.delete_cols(len(HEADERS) + 1, ws.max_column - len(HEADERS))

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    return wb

def _write_to_workbook(file_path: str, customer) -> None:
    wb = _get_or_create_workbook(file_path)
    ws = wb["Customer Leads"] if "Customer Leads" in wb.sheetnames else wb.active

    phone_display = format_phone_display(customer.whatsapp_number)
    first_ts = format_ist_timestamp(customer.first_contact_at)
    last_ts = format_ist_timestamp(customer.last_contact_at)
    
    # Priority: 1. Contact person / profile name, 2. Company name fallback
    contact_val = (customer.contact_person_name or "").strip()
    company = (customer.company_name or "").strip()
    if not contact_val or contact_val.lower() in ("customer", "none", ""):
        contact_val = company
    
    name = contact_val
    email = customer.email or ""
    company = customer.company_name or ""
    gst = customer.gst_number or ""
    address = customer.complete_address or ""
    reqs = customer.requirements_summary or ""

    # Find existing row by phone
    match_row = None
    first_empty = None

    for r in range(2, ws.max_row + 2):
        cv = ws.cell(row=r, column=4).value
        if cv:
            if str(cv).strip() == phone_display.strip():
                match_row = r
                break
        elif first_empty is None:
            first_empty = r

    align_c = Alignment(horizontal="center", vertical="center")
    align_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    if match_row:
        # Update Last Contact Date
        ws.cell(row=match_row, column=2, value=last_ts)
        if name: ws.cell(row=match_row, column=3, value=name)
        if email: ws.cell(row=match_row, column=5, value=email)
        if company: ws.cell(row=match_row, column=6, value=company)
        if gst: ws.cell(row=match_row, column=7, value=gst)
        if address: ws.cell(row=match_row, column=8, value=address)
        if reqs:
            existing_req = str(ws.cell(row=match_row, column=9).value or "").strip()
            if not existing_req:
                ws.cell(row=match_row, column=9, value=reqs)
            elif reqs.lower() not in existing_req.lower():
                ws.cell(row=match_row, column=9, value=f"{existing_req}\n{reqs}")
    else:
        new_row = first_empty if first_empty else (ws.max_row + 1)
        row_vals = [first_ts, last_ts, name, phone_display, email, company, gst, address, reqs]
        is_alt = (new_row % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=new_row, column=col_idx, value=val if val else None)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = ALT_FILL
            cell.alignment = align_c if col_idx in (1, 2, 4, 7) else align_l
        ws.row_dimensions[new_row].height = 28

    wb.save(file_path)

def _write_to_csv(file_path: str, customer) -> None:
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    rows = []
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    if not rows:
        rows = [HEADERS]

    phone_display = format_phone_display(customer.whatsapp_number)
    first_ts = format_ist_timestamp(customer.first_contact_at)
    last_ts = format_ist_timestamp(customer.last_contact_at)
    
    # Priority: 1. Contact person / profile name, 2. Company name fallback
    contact_val = (customer.contact_person_name or "").strip()
    company = (customer.company_name or "").strip()
    if not contact_val or contact_val.lower() in ("customer", "none", ""):
        contact_val = company
    
    name = contact_val
    email = customer.email or ""
    company = customer.company_name or ""
    gst = customer.gst_number or ""
    address = customer.complete_address or ""
    reqs = customer.requirements_summary or ""

    match_idx = None
    for i in range(1, len(rows)):
        if len(rows[i]) > 3 and rows[i][3] == phone_display:
            match_idx = i
            break

    if match_idx is not None:
        rows[match_idx][1] = last_ts
        if name and not rows[match_idx][2]: rows[match_idx][2] = name
        if email and not rows[match_idx][4]: rows[match_idx][4] = email
        if company and not rows[match_idx][5]: rows[match_idx][5] = company
        if gst and not rows[match_idx][6]: rows[match_idx][6] = gst
        if address and not rows[match_idx][7]: rows[match_idx][7] = address
        if reqs:
            cur_req = rows[match_idx][8] if len(rows[match_idx]) > 8 else ""
            if not cur_req:
                rows[match_idx][8] = reqs
            elif reqs.lower() not in cur_req.lower():
                rows[match_idx][8] = f"{cur_req} | {reqs}"
    else:
        rows.append([first_ts, last_ts, name, phone_display, email, company, gst, address, reqs])

    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)

_pending_lead_queue = []

def queue_pending_customer(customer):
    global _pending_lead_queue
    _pending_lead_queue.append(customer)

def flush_pending_excel_queue():
    global _pending_lead_queue
    if not _pending_lead_queue:
        return
    remaining = []
    with _excel_lock:
        for c in _pending_lead_queue:
            try:
                _write_to_workbook(settings.EXCEL_EXPORT_PATH, c)
            except PermissionError:
                remaining.append(c)
            except Exception:
                pass
    _pending_lead_queue = remaining
